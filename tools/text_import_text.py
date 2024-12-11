#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl import load_workbook
import subprocess
import sys
version = 'CR'
path = './'
path2 = ''

if len(sys.argv) > 1:
    version = sys.argv[1]

if version == 'CR':
    path = './pokecrystal_cn/'
else:
    path2 = 'xlsx/'

wb = load_workbook(path2+'text.xlsx', data_only=True)

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    
def RN(text):
    if text == None:
        return ''
    return text

class textblock:
    def __init__(self):
        self.dmap = ''
        self.omap = ''
        self.dlabel = ''
        self.olabel = ''
        self.ofname = ''
        self.oeomjp = ''
        self.oeomen = ''
        self.next_CR = False
        self.enlist = []
        self.jplist = []
        self.cnlist = []
        self.hint = []
        self.comment = []
        self.ctrl = []
        self.asm = ''
        self.version = ''

def printtb(cnlist, jplist, enlist):
    print("\n".join(["---CN---"] + cnlist))
    print("\n".join(["---JP---"] + jplist))
    print("\n".join(["---EN---"] + enlist))
    print("---==---")

ex_hint_dict = {}
with open('./tools/text_import_text_ex_hint.txt',encoding='utf-8') as f:
    for line in f:
        olabel, hint = line.strip('\n').split('\t')[:2]
        if olabel not in ex_hint_dict:
            ex_hint_dict[olabel] = []
        ex_hint_dict[olabel].append(hint)

ex_munt_dict = {}
with open('./tools/text_import_text_ex_munt.txt',encoding='utf-8') as f:
    for line in f:
        olabel = line.strip('\n').split('\t')[0]
        ex_munt_dict[olabel] = True

ex_sunt_dict = {}
with open('./tools/text_import_text_ex_sunt.txt',encoding='utf-8') as f:
    for line in f:
        olabel = line.strip('\n').split('\t')[0]
        ex_sunt_dict[olabel] = True

ex_epar_dict = {}
with open('./tools/text_import_text_ex_epar.txt',encoding='utf-8') as f:
    for line in f:
        olabel = line.strip('\n').split('\t')[0]
        ex_epar_dict[olabel] = True

ex_ctch_dict = {}
with open('./tools/text_import_text_ex_ctch.txt',encoding='utf-8') as f:
    for line in f:
        olabel = line.strip('\n').split('\t')[0]
        ex_ctch_dict[olabel] = True

ex_left_dict = {}
with open('./tools/text_import_text_ex_left.txt',encoding='utf-8') as f:
    for line in f:
        olabel = line.strip('\n').split('\t')[0]
        ex_left_dict[olabel] = True

def get_textdata():
    ws = wb['标']
    tb_dict = dict()
    tb_dict_version_not_matched = dict()
    for wbi in range(2, ws.max_row):
        currentVer = RN(ws.cell(row = wbi, column = 9).value).strip()
        tb = textblock()
        if ws.cell(row = wbi, column = 1).value is not None:
            tb.dmap = ws.cell(row = wbi, column = 1).value
        if ws.cell(row = wbi, column = 2).value is not None:
            tb.omap = ws.cell(row = wbi, column = 2).value
        if ws.cell(row = wbi, column = 3).value is not None:
            tb.dlabel = ws.cell(row = wbi, column = 3).value
        tb.olabel = ws.cell(row = wbi, column = 4).value
        tb.ofname = ws.cell(row = wbi, column = 5).value
        tb.oeomjp = ws.cell(row = wbi, column = 6).value
        tb.oeomen = ws.cell(row = wbi, column = 7).value
        tb.version = currentVer
        if currentVer == '' or currentVer == version:
            tb_dict[tb.olabel] = tb
        else:
            tb_dict_version_not_matched[tb.olabel] = tb

    sheetRange = 9
    if version != "CR":
        sheetRange = 10
    for i in range(sheetRange):
        ws = wb['文' + str(i + 1)]
        enlist = []
        jplist = []
        cnlist = []
        hint = []
        comment = []
        ctrl = []
        olabel = ''
        lastVer = ''
        for wbi in range(1, ws.max_row + 1):
            cen = ws.cell(row = wbi, column = 1).value
            if cen is None: cen = ''
            cjp = ws.cell(row = wbi, column = 3).value
            if cjp is None: cjp = ''
            ccn = ws.cell(row = wbi, column = 5).value
            if ccn is None: ccn = ''
            cht = ws.cell(row = wbi, column = 8).value
            if cht is None: cht = ''
            ccm = ws.cell(row = wbi, column = 9).value
            if ccm is None: ccm = ''
            ctr = ws.cell(row = wbi, column = 10).value
            if ctr is None: ctr = ''
            ver = ws.cell(row = wbi, column = 11).value
            if ver is None: ver = ''
            if '英文' in cen or '结束' in cen:
                if olabel != '':
                    if olabel in tb_dict_version_not_matched and not olabel in tb_dict:
                        print(f'olabel {olabel} 的版本不匹配 {tb_dict_version_not_matched[olabel].version} {version}')
                        olabel = ctr
                        lastVer = ver
                        enlist = []
                        jplist = []
                        cnlist = []
                        hint = []
                        comment = []
                        ctrl = []
                        continue
                    if lastVer != '' and lastVer != version:
                        print('L2: 当前文本版本不匹配，ignoring...')
                        print(olabel)
                        # print(ctr)
                        olabel = ctr
                        lastVer = ver
                        enlist = []
                        jplist = []
                        cnlist = []
                        hint = []
                        comment = []
                        ctrl = []
                        continue

                    enlist_end_cnt = 0
                    jplist_end_cnt = 0
                    while len(enlist) > 0 and enlist[-1] == '': 
                        enlist.pop()
                        enlist_end_cnt += 1
                    while len(jplist) > 0 and jplist[-1] == '': 
                        jplist.pop()
                        jplist_end_cnt += 1
                    while len(cnlist) > 0 and cnlist[-1] == '': cnlist.pop()
                    if enlist_end_cnt > 2 and jplist_end_cnt > 2 and olabel not in ex_epar_dict:
                        print(olabel, 'END WITH EMPTY PARA FOUND!')
                        printtb(cnlist, jplist, enlist)
                    if len(ctrl) > 0 and ctrl[0] == 'LINE_CR':
                        tb_dict[olabel].next_CR = True
                        ctrl.pop(0)
                    # if olabel == 'anata_msg_000_Kojindat':
                    #     ret1,codegit = subprocess.getstatusoutput('git -C ./pokecrystal_cn/ rev-parse --short HEAD')
                    #     ret2,textgit = subprocess.getstatusoutput('git -C . rev-parse --short HEAD')
                    #     print('INSERT GIT INFORMATION', ret1, codegit, ret2, textgit)
                    #     cnlist.insert(0, '感谢您参与《精灵宝可梦')
                    #     cnlist.insert(1, '水晶版》汉化版的测试！')
                    #     cnlist.insert(2, '')
                    #     cnlist.insert(3, '请在报告问题时')
                    #     cnlist.insert(4, '提供下一页的信息：')
                    #     cnlist.insert(5, '')
                    #     cnlist.insert(6, '代码：'+codegit)
                    #     cnlist.insert(7, '文本：'+textgit)
                    #     cnlist.insert(8, '')
                    #     cnlist.insert(9, '最后，请勿外传此测试')
                    #     cnlist.insert(10, 'ROM！下面开始游戏……')
                    #     cnlist.insert(11, '')
                    tb_dict[olabel].enlist = enlist
                    tb_dict[olabel].jplist = jplist
                    tb_dict[olabel].cnlist = cnlist
                    tb_dict[olabel].hint = hint
                    tb_dict[olabel].comment = comment
                    tb_dict[olabel].ctrl = ctrl
                    for hinttoken in hint:
                        if hinttoken == '仁': continue
                        if hinttoken == '真': continue
                        hintpass = False
                        for cnline in cnlist:
                            if hinttoken in cnline:
                                hintpass = True
                                break
                        if hintpass : continue
                        if olabel in ex_hint_dict:
                            if hinttoken in ex_hint_dict[olabel]:
                                hintpass = True
                                continue
                        print(olabel + ' HINT LOSS : ' + hinttoken)
                        print(''.join(cnlist))
                        
                    for jpline in jplist:
                        if '<USER>' in jpline:
                            hintpass = False
                            for cnline in cnlist:
                                if '<USER>' in cnline:
                                    hintpass = True
                                    break
                            if not hintpass :
                                print(olabel + ' HINT LOSS : ' + '<USER>')
                        if '<TARGET>' in jpline:
                            hintpass = False
                            for cnline in cnlist:
                                if '<TARGET>' in cnline:
                                    hintpass = True
                                    break
                            if not hintpass :
                                print(olabel + ' HINT LOSS : ' + '<TARGET>')

                if '英文' in cen:
                    olabel = ctr
                    lastVer = ver
                    enlist = []
                    jplist = []
                    cnlist = []
                    hint = []
                    comment = []
                    ctrl = []
            else:
                enlist.append(cen)
                jplist.append(cjp)
                cnlist.append(ccn)
                if cjp != '' and cjp == ccn and olabel not in ex_munt_dict:
                    print(olabel, 'MAYBE UNTRANSLATE:', cjp)
                if cht != '' :
                    for chttoken in cht.split(' '):
                        hint.append(chttoken.split(':')[1])
                if ccm != '' : comment.append(ccm)
                if ctr != '' : ctrl.append(ctr)
                if ctr == '…' :
                    print(olabel, 'WARN')
    return tb_dict
        


def get_asmfile_set():
    ws = wb['标']
    asmfile_set = set()
    for wbi in range(2, ws.max_row):
        currentVer = RN(ws.cell(row = wbi, column = 9).value).strip()
        if ws.cell(row = wbi, column = 1).value is not None:
            if currentVer == '' or currentVer == version:
                asmfile_set.add(ws.cell(row = wbi, column = 1).value)
            # else:
            #     print('L3: 当前版本不匹配，ignoring...')
            #     print(currentVer)
            #     print(version)
            #     print(RN(ws.cell(row = wbi, column = 1).value))
    return asmfile_set

def get_asmfile_data():
    asmfile_data = dict()
    for asmfile_name in get_asmfile_set():
        try:
            with open(path + asmfile_name,encoding='utf-8') as f:
                asmfile_data[asmfile_name] = f.readlines()
        except:
            print(f'ignoring {asmfile_name}')
    return asmfile_data


oddict = {}
spoddict = {}
with open('./tools/text_import_text_odctrl.txt',encoding='utf-8') as f:
    for line in f:
        od, fkname = line.strip('\n').split('\t')
        oddict[od] = fkname

with open('./tools/text_import_text_spodctrl.txt',encoding='utf-8') as f:
    for line in f:
        olabel, od, odtext = line.strip('\n').split('\t')[:3]
        if olabel not in spoddict:
            spoddict[olabel] = {}
        spoddict[olabel][od] = odtext

def length_check(tb):
    for line in tb.cnlist:
        # replace all name
        length_more = 0
        line = line.replace('—','―')
        line = line.replace('·', '・')
        line = line.replace('<……>', '..')
        line = line.replace('…', '.')
        line = line.replace('¥', 'Y')
        line = line.replace('<PLAYER>', 'PLAYERN')
        line = line.replace('<PLAY_G>', 'PLAYERN')
        line = line.replace('<RIVAL>',  'RIVALNM')
        line = line.replace('<USER>', '敌人的ABCDEFG')
        line = line.replace('<TARGET>', '敌人的ABCDEFG')
        # line = line.replace('<ENEMY>', '宝可梦训练家 CARINEY')
        line = line.replace('<ENEMY>', '宝可梦训练家CARINEY')
        line = line.replace('<SCROLL>', '')
        line = line.replace('\'s', 'S')
        for i, od in enumerate(tb.ctrl):
            if od == 'LINE_CR': 
                continue
            if od == 'text_low':
                length_more += 18
            if tb.olabel in spoddict and od in spoddict[tb.olabel]:
                line = line.replace('【' + str(i) + '】', spoddict[tb.olabel][od])
            else:
                line = line.replace('【' + str(i) + '】', oddict[od])
        if '<' in line: length = 180
        else: length = 0
        cnr = False
        for char in line:
            try:
                if char == '啰':
                    clen = 2
                else:
                    clen = len(char.encode(encoding='GB2312'))
            except:
                print(char, 'is not suppport', line, tb.dlabel)
                clen = 1
            if clen == 2:
                # chinese
                if not cnr :
                    length += 2
                    cnr = True
                else:
                    length += 1
                    cnr = False
            else:
                length += 1
                cnr = False
        if length > 18 + length_more:
            print("OVERF", length, tb.olabel, tb.dlabel, line)
            for od in tb.ctrl:
                if tb.olabel in spoddict and od in spoddict[tb.olabel]:
                    odtext = spoddict[tb.olabel][od]
                else:
                    odtext = oddict[od]
                print("ODTEXT\t" + tb.olabel + "\t" + od + "\t" + odtext)

def make_asm(tb):
    # asm_mk = tb.dlabel + ':\n'
    asm_mk = ''
    linec = 0
    parac = 0
    for line in tb.cnlist:
        if line == '':
            if linec == 0 and parac == 0:
                pass
            else:
                linec = 0
                parac += 1
                continue
        if linec == 0 and parac == 0:
            asm_mk += '\ttext "' + line + '"\n'
        elif linec == 0 and parac >= 1:
            asm_mk += '\n\tpara "' + line + '"\n'
        elif linec == 1:
            if tb.next_CR:
                asm_mk += '\tnext "' + line + '"\n'
            else:
                asm_mk += '\tline "' + line + '"\n'
        elif linec >= 2:
            if tb.next_CR:
                asm_mk += '\tnext "' + line + '"\n'
            else:
                asm_mk += '\tcont "' + line + '"\n'
        linec += 1
    for ctrli in range(len(tb.ctrl)):
        asm_mk = asm_mk.replace('【'+str(ctrli)+'】', '@"\n\t' + tb.ctrl[ctrli] + '\n\ttext "')
    asm_mk = asm_mk.replace('|', '')
    asm_mk = asm_mk.replace('text ""', "text_start")
    asm_mk = asm_mk.replace('\ttext "@"\n', '')
    if tb.oeomjp == 'EOMeom':
        asm_mk += '\tdone\n\n'
    elif tb.oeomjp == 'EOMwaiteom':
        asm_mk += '\tprompt\n\n'
    elif tb.oeomjp == 'EOM':
        asm_mk += '\ttext_end\n\n'
    elif tb.oeomjp == 'EOM^2':
        # asm_mk += '\ttext_end\n\n'
        asm_mk += '\ttext_end\n\n\ttext_end ; unreferenced\n\n'
    else:
        raise(Exception(tb.oeomjp + tb.olabel))
    # asm_mk = asm_mk.replace('\n\ttext_start\n\tdone\n', '\n\tdone\n')
    # asm_mk = asm_mk.replace('\n\ttext_start\n\tprompt\n', '\n\tprompt\n')
    asm_mk = asm_mk.replace('\n\ttext_start\n\ttext_end\n', '\n\ttext_end\n')
    asm_mk = asm_mk.replace('"\n\ttext_end\n\n\ttext_end ; unreferenced', '@"\n\ttext_end')
    asm_mk_list = asm_mk.splitlines()
    tst = False
    for line in asm_mk_list:
        if line == '':
            continue
        if tst == False:
            if 'text_start' in line:
                tst = True
        else:
            if 'para' in line or 'line' in line or 'cont' in line:
                pass
            elif 'done' in line or 'prompt' in line or 'text_end' in line:
                if tb.olabel not in ex_ctch_dict:
                    print('CATCH-', line, tb.olabel)
                    printtb(tb.cnlist, tb.jplist, tb.enlist)
            else:
                print('ERROR-', line, tb.olabel)
                printtb(tb.cnlist, tb.jplist, tb.enlist)
            tst = False
    if asm_mk == '\tdone\n\n':
        asm_mk = '\ttext_start\n' + asm_mk
    # print(asm_mk)
    # if 'text "' not in asm_mk and 'text_start' not in asm_mk:
    #     print('STRANGE', tb.dlabel)
    #     print(''.join(tb.jplist))
    #     print('----')
    #     print(''.join(tb.cnlist))
    #     print('----')
    #     print(asm_mk)
    #     print('====')
    return asm_mk

def get_textasm(tb_dict):
    tb_asm_dict = dict()
    for tbn in tb_dict:
        tb = tb_dict[tbn]
        tb.asm = make_asm(tb)
        if tb.dlabel != '':
            if tb_asm_dict.get(tb.dmap) is None:
                tb_asm_dict[tb.dmap] = dict()
            if tb_asm_dict[tb.dmap].get(tb.dlabel) is not None:
                print('ERROR! ', tb.dlabel)
            else:
                tb_asm_dict[tb.dmap][tb.dlabel] = tb
    return tb_asm_dict

def replace_asm(asmfile_list, asmn):
    label_found_asm = set()
    opt_list = []
    state = 0
    extra_end = False
    strip_au = 0
    for line in asmfile_list:
        line_strip = line[:line.find(';')].strip()
        if len(line_strip) > 0:
            if line_strip[-1] == ':':
                label = line_strip.strip(':')
                if state == 0:
                    tb = tb_asm_dict[asmn].get(label)
                elif state == 1:
                    if extra_end: extra_end = False
                        # tb.asm += '\ttext_end ; unreferenced\n\n'
                    opt_list.append(tb.asm)
                    length_check(tb)
                    tb = tb_asm_dict[asmn].get(label)
                if tb is not None:
                    label_found_asm.add(label)
                    if len(tb.jplist) != len(tb.cnlist): trans = True
                    else:
                        trans = False
                        for i in range(len(tb.jplist)):
                            if tb.jplist[i] != tb.cnlist[i]: trans = True
                        if trans == False and tb.olabel not in ex_sunt_dict:
                            print('UNTRANS TEXT?', tb.olabel)
                            printtb(tb.cnlist, tb.jplist, tb.enlist)
                            trans = True
                    if trans: state = 1
                    else: state = 0
                else: state = 0
            elif state == 1:
                if line == '\ttext_end ; unreferenced\n':
                    extra_end = True
                if line[0] == '\t':
                    line = '\t; ' + line[1:]
                else:
                    if strip_au == 0 and line.strip() == 'if DEF(_CRYSTAL_AU)':
                        strip_au = 1
                    elif strip_au == 1 and line.strip() == 'else':
                        strip_au = 2
                    elif strip_au == 2 and line.strip() == 'endc':
                        strip_au = 0
                    else:
                        strip_au = 0
                        print("UNKSRC", line.strip())
                    line = '; ' + line
        opt_list.append(line)
    if state == 1:
        opt_list.append(tb.asm)
    # print(''.join(opt_list))
    with open(path + asmn, 'w',encoding='utf-8') as f:
        f.writelines(opt_list)
    return label_found_asm

label_found = set()

asmfile_data = get_asmfile_data()
tb_dict = get_textdata()
tb_asm_dict = get_textasm(tb_dict)

for asmn in asmfile_data:
    asmfile = asmfile_data[asmn]
    label_found = label_found.union(replace_asm(asmfile, asmn))

# print(label_found)
for label in tb_dict:
    if tb_dict[label].dlabel not in label_found and label not in ex_left_dict:
        # pass
        print('LABEL LEFT', 'D[ ', tb_dict[label].dlabel, ' ] O[ ', label,' ]')
